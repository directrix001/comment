"""
xml_updater.py

Core safety rule: NEVER run this workbook through openpyxl load+save as a whole.
openpyxl drops external-link definitions on save, which is exactly the link
breakage we must avoid. Instead we edit the raw OOXML parts directly:

  1. Re-point each externalLinkN.xml.rels Target at the new source file the
     user matched, so Excel finds the right file next time it opens.
  2. Refresh the cached values inside externalLinkN.xml (sheetDataSet) by
     reading the actual current values from that new source file, so the
     numbers are already correct even before Excel recalculates.
  3. Write any manual-input values the user typed in the UI straight into the
     relevant sheetN.xml cell's <v>, leaving every formula cell completely
     untouched.

All three are pure string/zip operations - the workbook is never re-serialized
by openpyxl, so nothing else in the file (styles, formulas, external link
structure) can be disturbed.
"""
import os
import re
import shutil
import zipfile
import urllib.parse
import openpyxl


def _win_path_to_file_uri(path):
    p = path.replace('\\', '/')
    if not p.startswith('/'):
        p = '/' + p
    return 'file://' + urllib.parse.quote(p, safe='/:()&+-_.,\'')


def relink_external_files(src_xlsx, dst_xlsx, index_to_newpath):
    """
    index_to_newpath: dict[int -> absolute path on disk of the replacement file]
    Only indices present in the dict are touched; everything else in the zip
    is copied through byte-for-byte.
    """
    if os.path.abspath(src_xlsx) == os.path.abspath(dst_xlsx):
        raise ValueError('dst_xlsx must differ from src_xlsx')

    shutil.copy2(src_xlsx, dst_xlsx)

    with zipfile.ZipFile(src_xlsx) as zin:
        names = zin.namelist()

    for idx, new_path in index_to_newpath.items():
        if not new_path:
            continue
        rels_name = f'xl/externalLinks/_rels/externalLink{idx}.xml.rels'
        if rels_name not in names:
            continue
        _rewrite_zip_entry(dst_xlsx, rels_name, lambda xml, np=new_path: _patch_rels_target(xml, np))


def _patch_rels_target(rels_xml, new_path):
    new_uri = _win_path_to_file_uri(new_path)
    def repl(m):
        return m.group(1) + new_uri + m.group(3)
    # replace every Target="..." that is an externalLinkPath relationship
    return re.sub(r'(Target=")([^"]*)(")', repl, rels_xml)


def refresh_cached_values(dst_xlsx, links, index_to_newpath, log=None):
    """
    For every relinked external link, open the *new* source file (read-only,
    values only) and overwrite the cached numbers inside externalLinkN.xml so
    the file already shows current numbers, without touching any formula.
    Failures on one link are swallowed (old cache just stays as-is - never
    fatal, never removes a link) and reported through `log` if provided.
    """
    log = log or (lambda msg: None)
    with zipfile.ZipFile(dst_xlsx) as zin:
        names = zin.namelist()

    for idx, new_path in index_to_newpath.items():
        if not new_path or not os.path.isfile(new_path):
            continue
        xml_name = f'xl/externalLinks/externalLink{idx}.xml'
        if xml_name not in names:
            continue
        try:
            wb_src = openpyxl.load_workbook(new_path, data_only=True, read_only=True)
        except Exception as e:
            log(f'[{idx}] could not open matched file for value refresh: {e}')
            continue

        def patch(xml, wb_src=wb_src, idx=idx):
            return _patch_external_cache(xml, wb_src, log, idx)

        try:
            _rewrite_zip_entry(dst_xlsx, xml_name, patch)
        except Exception as e:
            log(f'[{idx}] cache refresh failed, link path still updated: {e}')
        finally:
            wb_src.close()


def _patch_external_cache(xml, wb_src, log, idx):
    sheet_names = re.findall(r'<sheetName val="([^"]+)"/>', xml)

    def patch_sheetdata(m):
        sheet_id = int(m.group('sid'))
        body = m.group('body')
        if sheet_id >= len(sheet_names):
            return m.group(0)
        sname = sheet_names[sheet_id].replace('&amp;', '&')
        if sname not in wb_src.sheetnames:
            return m.group(0)
        ws = wb_src[sname]

        def patch_cell(cm):
            cell_ref = cm.group('ref')
            try:
                new_val = ws[cell_ref].value
            except Exception:
                return cm.group(0)
            if new_val is None:
                return cm.group(0)  # keep old cached value rather than blank it
            if isinstance(new_val, str):
                # cached external values in this file are numeric in practice;
                # skip strings to avoid corrupting a numeric <v> with text.
                return cm.group(0)
            return f'<cell r="{cell_ref}"><v>{new_val}</v></cell>'

        new_body = re.sub(r'<cell r="(?P<ref>[A-Z]+\d+)"[^>]*>(?:<v>[^<]*</v>)?</cell>', patch_cell, body)
        return m.group('head') + new_body + m.group('tail')

    pattern = re.compile(
        r'(?P<head><sheetData sheetId="(?P<sid>\d+)"[^>]*>)(?P<body>.*?)(?P<tail></sheetData>)',
        re.S,
    )
    return pattern.sub(patch_sheetdata, xml)


def write_manual_values(dst_xlsx, sheet_to_file, updates, log=None):
    """
    updates: list of (sheet_name, cell_ref, new_numeric_value)
    Groups by sheet file and rewrites only the targeted <v> for each cell,
    leaving every formula in that sheet (and every other sheet) untouched.
    """
    log = log or (lambda msg: None)
    by_sheet = {}
    for sheet_name, cell_ref, value in updates:
        sf = sheet_to_file.get(sheet_name)
        if not sf:
            log(f'Unknown sheet for manual update: {sheet_name}')
            continue
        by_sheet.setdefault(sf, []).append((cell_ref, value))

    for sf, cell_updates in by_sheet.items():
        def patch(xml, cell_updates=cell_updates):
            for cell_ref, value in cell_updates:
                m = re.search(r'<c r="%s"[^>]*>' % re.escape(cell_ref), xml)
                if not m:
                    continue
                tag_start, tag_end = m.span()
                close = xml.index('</c>', tag_end)
                open_tag = xml[tag_start:tag_end]
                if '<f' in xml[tag_end:close]:
                    fm = re.search(r'<f[^>]*(?:/>|>.*?</f>)', xml[tag_end:close], re.S)
                    formula_part = fm.group(0) if fm else ''
                    new_cell = f'{open_tag}{formula_part}<v>{value}</v></c>'
                else:
                    new_cell = f'{open_tag}<v>{value}</v></c>'
                xml = xml[:tag_start] + new_cell + xml[close + len('</c>'):]
            return xml
        _rewrite_zip_entry(dst_xlsx, sf, patch)


def _rewrite_zip_entry(zip_path, entry_name, transform_fn):
    """
    Replace one entry inside a zip file with transform_fn(current_text_content),
    rewriting the whole archive to a temp file and swapping it in (zipfile has
    no in-place edit). Every other entry is copied through unchanged and
    uncompressed/compressed identically to how zipfile normally writes it.
    """
    tmp_path = zip_path + '.tmp'
    with zipfile.ZipFile(zip_path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == entry_name:
                text = data.decode('utf-8', errors='replace')
                new_text = transform_fn(text)
                data = new_text.encode('utf-8')
            zout.writestr(item, data)
    os.replace(tmp_path, zip_path)


def sheet_name_to_file_map(xlsx_path):
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read('xl/workbook.xml').decode('utf-8', errors='replace')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
    rid_to_target = dict(re.findall(r'<Relationship Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))
    sheets = re.findall(r'<sheet name="([^"]*)"[^>]*r:id="(rId\d+)"', wb_xml)
    mapping = {}
    for name, rid in sheets:
        target = rid_to_target.get(rid)
        if target:
            norm = 'xl/' + target if not target.startswith('xl/') else target
            mapping[name.replace('&amp;', '&')] = norm
    return mapping
